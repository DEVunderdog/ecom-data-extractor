"""Extract Swagify header row (row 1) from the reference XLSX into JSON.

Run once at setup. Preserves column order and duplicate column names.
"""
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "swagify_reference.xlsx"
OUT = ROOT / "data" / "swagify_headers.json"


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            if cell is None:
                continue
            headers.append(str(cell).strip())
        break
    if len(headers) < 190:
        raise SystemExit(f"Refusing to write: header count {len(headers)} < 190")
    OUT.write_text(json.dumps(headers, ensure_ascii=False, indent=2), encoding="utf-8")
    total = len(headers)
    uniq = len(set(headers))
    print(f"Wrote {OUT} — {total} columns ({total - uniq} duplicates)")
    print("First 8:", headers[:8])
    print("Last 5:", headers[-5:])


if __name__ == "__main__":
    main()
